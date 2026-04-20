"""Generate sample_shifts.xlsx for PharmShift download template."""
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Color palette ──────────────────────────────────────────────────────────────
C_PURPLE_DARK  = "8B5CF6"
C_PURPLE_MED   = "7C3AED"
C_PURPLE_LIGHT = "EDE9FE"
C_PURPLE_XLIGHT= "F5F3FF"
C_TEAL_DARK    = "0891B2"
C_GREEN_DARK   = "059669"
C_GRAY_LIGHT   = "F3F4F6"
C_GRAY_TEXT    = "6B7280"

SHIFT_FILL = {
    "เช้า":     ("E8F9FA", "9FDCE0"),
    "บ่าย":     ("F3EDF8", "9E76B4"),
    "ดึก":      ("EEF0FF", "99ABFF"),
    "รุ่งอรุณ": ("FEF3DC", "FFCA72"),
    "multi":    ("FFF9E6", "F59E0B"),
}

# Classify a shift code → (type_key, display)
def classify(code: str) -> str:
    c = code.strip().lower()
    if c == "ด":
        return "ดึก"
    if c.startswith("ร"):
        return "รุ่งอรุณ"
    if c.startswith("บ") or c == "smc" or c.startswith("smc") or c.startswith("ext"):
        return "บ่าย"
    return "เช้า"

def get_fill(code: str):
    if "/" in code or "," in code:
        bg, bd = SHIFT_FILL["multi"]
    else:
        t = classify(code)
        bg, bd = SHIFT_FILL.get(t, ("FFFFFF", "CCCCCC"))
    return PatternFill("solid", start_color=bg, fgColor=bg), bd

def thin_border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def header_fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def make_sheet(role_label: str, tab_color: str, rows_data: list[dict]):
    """
    rows_data: list of {"nickname": str, "pha_id": str, "shifts": dict[int, str]}
    Column A = ชื่อเล่น (ignored by server, visual reference)
    Column B = pha_id   (used by server for user lookup)
    """
    ws = wb.create_sheet(title=role_label)
    ws.sheet_properties.tabColor = tab_color

    DAYS = 30  # April has 30 days

    # ── Row 1: header ─────────────────────────────────────────────────────────
    # A1 = role identifier (server reads this for role detection)
    ws.cell(1, 1, role_label)
    ws.cell(1, 1).font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    ws.cell(1, 1).fill = header_fill(C_PURPLE_DARK)
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")

    # B1 = pha_id header label
    ws.cell(1, 2, "pha_id")
    ws.cell(1, 2).font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    ws.cell(1, 2).fill = header_fill(C_PURPLE_DARK)
    ws.cell(1, 2).alignment = Alignment(horizontal="center", vertical="center")

    # C1–AF1 = days 1-30
    for d in range(1, DAYS + 1):
        col = d + 2
        cell = ws.cell(1, col, str(d))
        cell.font = Font(name="Calibri", bold=True, size=11, color="5B21B6")
        cell.fill = header_fill(C_PURPLE_LIGHT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border("C4B5FD")

    ws.row_dimensions[1].height = 28

    # ── Data rows ─────────────────────────────────────────────────────────────
    for r_idx, person in enumerate(rows_data, start=2):
        nickname = person["nickname"]
        pha_id   = person["pha_id"]
        shifts   = person["shifts"]  # {day: code}
        row_bg   = "FAF5FF" if r_idx % 2 == 0 else "FFFFFF"

        # A: ชื่อเล่น (ignored by server — visual reference only)
        ws.cell(r_idx, 1, nickname)
        ws.cell(r_idx, 1).font = Font(name="Calibri", size=10, color=C_GRAY_TEXT)
        ws.cell(r_idx, 1).fill = PatternFill("solid", start_color=C_GRAY_LIGHT, fgColor=C_GRAY_LIGHT)
        ws.cell(r_idx, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(r_idx, 1).border = thin_border("E5E7EB")

        # B: pha_id (server uses this for user lookup)
        ws.cell(r_idx, 2, pha_id)
        ws.cell(r_idx, 2).font = Font(name="Calibri", bold=True, size=11, color="4C1D95")
        ws.cell(r_idx, 2).fill = PatternFill("solid", start_color=C_PURPLE_XLIGHT, fgColor=C_PURPLE_XLIGHT)
        ws.cell(r_idx, 2).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.cell(r_idx, 2).border = thin_border("C4B5FD")

        # C–AF: shift cells
        for d in range(1, DAYS + 1):
            col = d + 2
            code = shifts.get(d, "")
            cell = ws.cell(r_idx, col, code)
            cell.font = Font(name="Calibri", size=10, color="1F2937")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if code:
                fill, border_color = get_fill(code)
                cell.fill = fill
                cell.border = thin_border(border_color)
            else:
                cell.fill = PatternFill("solid", start_color=row_bg, fgColor=row_bg)
                cell.border = thin_border("E5E7EB")

        ws.row_dimensions[r_idx].height = 22

    # ── Legend ────────────────────────────────────────────────────────────────
    legend_row = len(rows_data) + 4
    ws.cell(legend_row, 1, "ตัวอย่างสีเวร:")
    ws.cell(legend_row, 1).font = Font(name="Calibri", bold=True, size=10, color="374151")

    legend_items = [
        ("เช้า",     "E8F9FA", "9FDCE0", "teal"),
        ("บ่าย",     "F3EDF8", "9E76B4", "purple"),
        ("ดึก",      "EEF0FF", "99ABFF", "indigo"),
        ("รุ่งอรุณ", "FEF3DC", "FFCA72", "amber"),
    ]
    for i, (label, bg, bd, _) in enumerate(legend_items):
        col = i * 4 + 2
        ws.cell(legend_row, col, label)
        ws.cell(legend_row, col).font = Font(name="Calibri", bold=True, size=10, color="1F2937")
        ws.cell(legend_row, col).fill = PatternFill("solid", start_color=bg, fgColor=bg)
        ws.cell(legend_row, col).alignment = Alignment(horizontal="center")
        ws.cell(legend_row, col).border = thin_border(bd)
        ws.merge_cells(start_row=legend_row, start_column=col,
                       end_row=legend_row, end_column=col + 2)

    ws.row_dimensions[legend_row].height = 20

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 12   # ชื่อเล่น
    ws.column_dimensions["B"].width = 12   # pha_id
    for d in range(1, DAYS + 1):
        ws.column_dimensions[get_column_letter(d + 2)].width = 5.5

    # ── Freeze panes after column B ───────────────────────────────────────────
    ws.freeze_panes = "C2"

    return ws


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1: เภสัชกร
# ══════════════════════════════════════════════════════════════════════════════
pha_rows = [
    {"nickname": "สมชาย", "pha_id": "pha001", "shifts": {
        1:"d", 2:"d", 3:"s", 4:"e", 5:"d",
        7:"รo", 8:"รo", 9:"d", 10:"s",
        12:"e", 14:"c", 15:"ด",
        17:"s", 18:"d", 19:"บm",
        20:"ext", 21:"e", 22:"c",
        25:"smc", 26:"d", 27:"e",
        28:"รe", 29:"s",
    }},
    {"nickname": "สมหญิง", "pha_id": "pha002", "shifts": {
        1:"s", 2:"s", 3:"e", 4:"e", 5:"c",
        7:"c", 8:"บm", 9:"s",
        11:"e", 12:"รe", 13:"d",
        16:"บe", 17:"s",
        20:"c", 21:"e", 22:"ด",
        24:"d", 25:"s",
        27:"รo", 28:"c",
    }},
    {"nickname": "วิภา", "pha_id": "pha003", "shifts": {
        2:"c", 3:"c", 4:"d", 5:"s", 6:"e",
        8:"e", 9:"รh", 10:"d",
        12:"s", 13:"c",
        16:"ext", 17:"e",
        19:"d", 20:"s",
        23:"ch", 24:"e",
        26:"c", 27:"d",
        29:"รo", 30:"s",
    }},
    {"nickname": "ธนกร", "pha_id": "pha004", "shifts": {
        1:"e", 2:"e", 3:"บe", 4:"d", 5:"c",
        7:"s", 8:"e", 9:"d",
        11:"smc", 12:"e",
        14:"c", 15:"s",
        19:"ด", 20:"e",
        22:"d", 23:"c",
        25:"s", 26:"e",
        28:"รo", 29:"d",
    }},
    {"nickname": "พิมพ์", "pha_id": "pha005", "shifts": {
        1:"บm", 2:"บe", 3:"smc", 4:"d",
        6:"e", 7:"s", 8:"c",
        11:"d", 12:"e",
        14:"รo/ด", 15:"s",
        17:"c", 18:"d",
        21:"ext", 22:"e",
        24:"s", 25:"บm",
        27:"d", 28:"c",
        30:"รe",
    }},
]
make_sheet("เภสัช", "7C3AED", pha_rows)


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 2: จพง
# ══════════════════════════════════════════════════════════════════════════════
jpg_rows = [
    {"nickname": "มานะ", "pha_id": "pha006", "shifts": {
        1:"m1", 2:"m2", 3:"s1", 4:"m3", 5:"m1",
        7:"m2", 8:"บm1", 9:"s2",
        11:"m1", 12:"m3", 13:"s1",
        14:"ด", 15:"m2",
        17:"บm2", 18:"m1",
        20:"smc1", 21:"m3",
        23:"s2", 24:"m1",
        25:"รo", 26:"m2",
        28:"s1", 29:"m3",
    }},
    {"nickname": "สุดา", "pha_id": "pha007", "shifts": {
        1:"s2", 2:"m1", 3:"m2", 4:"บe", 5:"s1",
        7:"m3", 8:"m1", 9:"s2",
        10:"รe", 11:"m2",
        13:"s1", 14:"m3",
        16:"m1", 17:"s2",
        18:"ด", 19:"m1",
        21:"บm1", 22:"m3",
        24:"บm2", 25:"s1",
        27:"m2", 28:"รh",
        30:"ด",
    }},
    {"nickname": "กนก", "pha_id": "pha008", "shifts": {
        2:"m3", 3:"s1", 4:"m1", 5:"บm1", 6:"m2",
        8:"s2", 9:"m3", 10:"m1",
        12:"smc2", 13:"m2",
        15:"s1", 16:"m3",
        19:"รh", 20:"m1",
        22:"m2", 23:"s2",
        25:"m3", 26:"บm2",
        27:"ด", 28:"m1",
        30:"m2",
    }},
    {"nickname": "อรทัย", "pha_id": "pha009", "shifts": {
        1:"บm2", 2:"บe", 3:"m3", 4:"s2",
        6:"m1", 7:"m2", 8:"s1",
        10:"m3", 11:"รo",
        13:"m1", 14:"s2",
        16:"ด", 17:"m2",
        19:"m3", 20:"s1",
        22:"smc1", 23:"m1",
        25:"m2", 26:"s2",
        28:"m3", 29:"รe",
    }},
]
make_sheet("จพง", "0891B2", jpg_rows)


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 3: จนท
# ══════════════════════════════════════════════════════════════════════════════
jnt_rows = [
    {"nickname": "ประยุทธ", "pha_id": "pha010", "shifts": {
        1:"m1", 2:"m2", 3:"s1", 4:"m3", 5:"m4",
        7:"m1", 8:"บm", 9:"m2",
        11:"s2", 12:"m3", 13:"ด",
        15:"m1", 16:"s3",
        18:"m4", 19:"m2",
        20:"ext1", 21:"m1",
        23:"s1", 24:"m3",
        26:"ส1", 27:"m2",
        29:"s2", 30:"m4",
    }},
    {"nickname": "นภา", "pha_id": "pha011", "shifts": {
        1:"s2", 2:"m1", 3:"บe1", 4:"s3", 5:"m2",
        7:"m3", 8:"s1", 9:"รo1",
        11:"m4", 12:"m2",
        14:"s2", 15:"m1",
        17:"ด", 18:"m3",
        20:"s3", 21:"m4",
        23:"smc1", 24:"m1",
        25:"s1", 26:"m2",
        28:"ส2", 29:"m3",
    }},
    {"nickname": "สาวิตรี", "pha_id": "pha012", "shifts": {
        2:"m3", 3:"ext2", 4:"m1", 5:"s1", 6:"บe2",
        8:"m2", 9:"m4", 10:"s3",
        11:"รe", 12:"m1",
        14:"m3", 15:"s2",
        17:"m4", 18:"ส3",
        20:"m1", 21:"s1",
        23:"m2", 24:"ext1",
        25:"ด", 26:"m3",
        28:"s2", 29:"m4",
    }},
    {"nickname": "รัตนา", "pha_id": "pha013", "shifts": {
        1:"บm", 2:"รo2", 3:"m4", 4:"s2",
        6:"m1", 7:"m3", 8:"s1",
        10:"smc2", 11:"m2",
        13:"s3", 14:"m4",
        15:"ด", 16:"m1",
        18:"m3", 19:"s2",
        22:"ส4", 23:"m1",
        25:"m2", 26:"s1",
        27:"ext1", 28:"m4",
        29:"m3", 30:"รo1",
    }},
]
make_sheet("จนท", "059669", jnt_rows)


# Remove the default blank sheet
del wb["Sheet"]

# Save
out_path = "/Users/codex074/Desktop/My-Web-App/pharmshift/public/sample_shifts.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
